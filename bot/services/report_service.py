from sqlalchemy import select, func, and_
from datetime import date, datetime, timedelta
from typing import Tuple, Optional, Dict, List, Any
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from bot.database import async_session_maker
from bot.models.transaction import Transaction
from bot.models.category import Category
from bot.models.account import Account
from bot.models.budget import Budget
from bot.models.goal import Goal
from bot.services.chart_service import chart_service

# Ensure exports directory exists
EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

class ReportService:
    async def get_summary_stats(self, user_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
        """Calculates total incomes, total expenses, net savings, and category distribution (in UZS) for a date range."""
        async with async_session_maker() as session:
            # 1. Total Incomes and Expenses
            txs_res = await session.execute(
                select(Transaction.type, func.sum(Transaction.amount_uzs)).where(
                    Transaction.user_id == user_id,
                    Transaction.date >= start_date,
                    Transaction.date <= end_date
                ).group_by(Transaction.type)
            )
            totals = {row[0]: float(row[1] or 0.0) for row in txs_res.all()}
            
            incomes = totals.get("income", 0.0)
            expenses = totals.get("expense", 0.0)
            net_savings = incomes - expenses
            savings_pct = (net_savings / incomes * 100) if incomes > 0 else 0.0

            # 2. Category distribution for expenses
            cat_res = await session.execute(
                select(Category.name, Category.icon, func.sum(Transaction.amount_uzs)).select_from(Transaction)
                .join(Category, Transaction.category_id == Category.id)
                .where(
                    Transaction.user_id == user_id,
                    Transaction.type == "expense",
                    Transaction.date >= start_date,
                    Transaction.date <= end_date
                ).group_by(Category.name, Category.icon)
            )
            category_distribution = {}
            for row in cat_res.all():
                category_distribution[f"{row[1]} {row[0]}"] = float(row[2] or 0.0)

            # 3. Top-5 expenses
            top_res = await session.execute(
                select(Category.name, Category.icon, Transaction.amount, Transaction.currency, Transaction.description, Transaction.date)
                .select_from(Transaction)
                .join(Category, Transaction.category_id == Category.id)
                .where(
                    Transaction.user_id == user_id,
                    Transaction.type == "expense",
                    Transaction.date >= start_date,
                    Transaction.date <= end_date
                ).order_by(Transaction.amount_uzs.desc()).limit(5)
            )
            top_expenses = []
            for row in top_res.all():
                top_expenses.append({
                    "category": f"{row[1]} {row[0]}",
                    "amount": float(row[2]),
                    "currency": row[3],
                    "desc": row[4] or "",
                    "date": row[5].strftime("%d.%m")
                })

            return {
                "incomes": incomes,
                "expenses": expenses,
                "savings": net_savings,
                "savings_pct": savings_pct,
                "categories": category_distribution,
                "top_expenses": top_expenses
            }

    async def generate_daily_report(self, user_id: int, target_date: date = None) -> Tuple[str, Optional[str]]:
        """Generates a text daily report and returns (text, chart_path)."""
        target_date = target_date or date.today()
        yesterday = target_date - timedelta(days=1)
        
        stats = await self.get_summary_stats(user_id, target_date, target_date)
        yesterday_stats = await self.get_summary_stats(user_id, yesterday, yesterday)
        
        incomes = stats["incomes"]
        expenses = stats["expenses"]
        balance = incomes - expenses
        
        # Compare with yesterday
        yesterday_exp = yesterday_stats["expenses"]
        if yesterday_exp > 0:
            diff = ((yesterday_exp - expenses) / yesterday_exp) * 100
            diff_text = f"+{round(abs(diff), 1)}% kamroq" if diff >= 0 else f"+{round(abs(diff), 1)}% ko'proq"
        else:
            diff_text = "Taqqoslash uchun ma'lumot yo'q"
            
        # Format text
        text = f"📊 **BUGUNGI HISOBOT — {target_date.strftime('%d.%m.%Y')}**\n"
        text += "━━━━━━━━━━━━━━━━━━━━━━\n"
        text += f"💰 **Kirim:** {incomes:,.0f} so'm\n"
        text += f"💸 **Chiqim:** {expenses:,.0f} so'm\n"
        text += f"💵 **Balans:** {balance:,.0f} so'm\n"
        text += "━━━━━━━━━━━━━━━━━━━━━━\n"
        
        if stats["categories"]:
            text += "📌 **KATEGORIYALAR:**\n"
            for cat_name, amt in stats["categories"].items():
                pct = (amt / expenses * 100) if expenses > 0 else 0
                text += f"{cat_name}: {amt:,.0f} so'm ({pct:.0f}%)\n"
            text += "━━━━━━━━━━━━━━━━━━━━━━\n"
            
        text += f"📈 Kecha bilan: {diff_text}\n"
        
        # Generate chart if there are expenses
        chart_path = None
        if stats["categories"]:
            chart_path = chart_service.generate_expense_pie_chart(user_id, stats["categories"])
            
        return text, chart_path

    async def generate_weekly_report(self, user_id: int) -> Tuple[str, Optional[str]]:
        """Generates a weekly report for the last 7 days."""
        end_date = date.today()
        start_date = end_date - timedelta(days=6)
        
        stats = await self.get_summary_stats(user_id, start_date, end_date)
        
        text = f"📅 **HAFTALIK HISOBOT ({start_date.strftime('%d.%m')} - {end_date.strftime('%d.%m')})**\n"
        text += "━━━━━━━━━━━━━━━━━━━━━━\n"
        text += f"💰 **Haftalik Kirim:** {stats['incomes']:,.0f} so'm\n"
        text += f"💸 **Haftalik Chiqim:** {stats['expenses']:,.0f} so'm\n"
        text += f"💎 **Tejov:** {stats['savings']:,.0f} so'm ({stats['savings_pct']:.0f}%)\n"
        text += "━━━━━━━━━━━━━━━━━━━━━━\n"
        
        if stats["top_expenses"]:
            text += "🏆 **ENG KATTA XARAJATLAR (TOP-5):**\n"
            for idx, item in enumerate(stats["top_expenses"], 1):
                desc_str = f" ({item['desc']})" if item["desc"] else ""
                text += f"{idx}. {item['category']} — {item['amount']:,.0f} {item['currency']}{desc_str}\n"
            text += "━━━━━━━━━━━━━━━━━━━━━━\n"
            
        # Draw trend line chart
        chart_path = None
        # We need trend data for each day in last 7 days
        days_list = []
        amounts_list = []
        async with async_session_maker() as session:
            for i in range(7):
                day = start_date + timedelta(days=i)
                res = await session.execute(
                    select(func.sum(Transaction.amount_uzs)).where(
                        Transaction.user_id == user_id,
                        Transaction.type == "expense",
                        Transaction.date == day
                    )
                )
                amt = float(res.scalar() or 0.0)
                days_list.append(day.strftime("%a")) # Mon, Tue, etc.
                amounts_list.append(amt)
                
        if sum(amounts_list) > 0:
            chart_path = chart_service.generate_trend_chart(user_id, days_list, amounts_list)
            
        return text, chart_path

    async def generate_monthly_report(self, user_id: int) -> Tuple[str, Optional[str]]:
        """Generates a monthly report for the current month."""
        today = date.today()
        start_date = date(today.year, today.month, 1)
        
        stats = await self.get_summary_stats(user_id, start_date, today)
        
        text = f"📅 **{today.strftime('%B %Y').upper()} — OYLIK HISOBOT**\n"
        text += "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        text += f"💰 **Jami kirim:**  {stats['incomes']:,.0f} so'm\n"
        text += f"💸 **Jami chiqim:** {stats['expenses']:,.0f} so'm\n"
        text += f"💎 **Tejaldi:**      {stats['savings']:,.0f} so'm\n"
        text += f"📊 **Tejash %:**     {stats['savings_pct']:.0f}%\n"
        text += "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        
        if stats["categories"]:
            text += "🏆 **ENG KO'P XARAJAT:**\n"
            sorted_cats = sorted(stats["categories"].items(), key=lambda x: x[1], reverse=True)[:5]
            for idx, (cat_name, amt) in enumerate(sorted_cats, 1):
                pct = (amt / stats['expenses'] * 100) if stats['expenses'] > 0 else 0
                text += f"{idx}. {cat_name}: {amt:,.0f} so'm ({pct:.0f}%)\n"
            text += "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"

        # Check goals progress
        async with async_session_maker() as session:
            goals_res = await session.execute(
                select(Goal).where(Goal.user_id == user_id, Goal.is_achieved == False)
            )
            goals = goals_res.scalars().all()
            if goals:
                text += "🎯 **MAQSADLAR PROGRESSI:**\n"
                for g in goals:
                    pct = (float(g.current_amount) / float(g.target_amount) * 100) if g.target_amount > 0 else 0
                    text += f"🎯 {g.name}: {pct:.1f}% ({g.current_amount:,.0f} / {g.target_amount:,.0f} {g.currency})\n"
                text += "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"

        chart_path = None
        if stats["categories"]:
            chart_path = chart_service.generate_expense_pie_chart(user_id, stats["categories"])
            
        return text, chart_path

    async def export_to_excel(self, user_id: int, start_date: date, end_date: date) -> Optional[str]:
        """
        Exports all transactions in the specified period to a beautiful Excel sheet.
        Returns the absolute file path of the excel file.
        """
        async with async_session_maker() as session:
            # Join with Account and Category
            tx_res = await session.execute(
                select(Transaction, Account.name, Category.name).select_from(Transaction)
                .outerjoin(Account, Transaction.account_id == Account.id)
                .outerjoin(Category, Transaction.category_id == Category.id)
                .where(
                    Transaction.user_id == user_id,
                    Transaction.date >= start_date,
                    Transaction.date <= end_date
                ).order_by(Transaction.date.desc(), Transaction.id.desc())
            )
            rows = tx_res.all()

            if not rows:
                return None

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tranzaksiyalar"
            
            # Enable gridlines
            ws.views.sheetView[0].showGridLines = True
            
            # Header Row
            headers = ["ID", "Sana", "Turi", "Hisob", "Kategoriya", "Miqdor", "Valyuta", "Miqdor (UZS)", "Tavsif"]
            ws.append(headers)
            
            # Format header
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Append transactions
            for tx, acc_name, cat_name in rows:
                ws.append([
                    tx.id,
                    tx.date.strftime("%Y-%m-%d"),
                    "Kirim" if tx.type == "income" else "Chiqim" if tx.type == "expense" else "O'tkazma",
                    acc_name or "Noma'lum",
                    cat_name or "Noma'lum",
                    float(tx.amount),
                    tx.currency,
                    float(tx.amount_uzs),
                    tx.description or ""
                ])

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            file_path = os.path.join(EXPORTS_DIR, f"finance_report_{user_id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx")
            wb.save(file_path)
            return file_path

report_service = ReportService()
